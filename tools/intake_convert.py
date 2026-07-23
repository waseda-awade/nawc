#!/usr/bin/env python3
"""intake_convert.py - editor-side conversion of raw submitted texts to corpus format.

Converts a directory of .txt/.docx files, or a spreadsheet with one text per row,
into NFC-normalized UTF-8 text files plus a manifest skeleton and a processing log
with SHA-256 checksums of every source file (chain of custody).

Usage:
  # Directory of files (one essay per .txt or .docx file):
  python tools/intake_convert.py --input raw_submission/ --outdir data/kyoto-argumentative-2026 --prefix kyo26

  # Spreadsheet (one essay per row):
  python tools/intake_convert.py --input essays.xlsx --outdir data/kyoto-argumentative-2026 \
      --prefix kyo26 --text-col "essay" [--id-col "student"] [--sheet "Sheet1"]

  # Spreadsheet with essays AND scores in the same file:
  python tools/intake_convert.py --input essays.xlsx --outdir data/awade-2025 --prefix awade \
      --text-col "essay" --id-col "student" --score-cols "rater1_score,rater2_score"

  # Individual files, with scores in a SEPARATE file matched by original filename:
  python tools/intake_convert.py --input raw_submission/ --outdir data/awade-2025 --prefix awade \
      --scores-file scores.csv --scores-key-col "filename" --scores-value-cols "score"

Outputs under --outdir:
  texts/<prefix>-NNNN.txt      one per essay, UTF-8, NFC, LF line endings
  manifest_skeleton.csv        text_id + word_count (+ score, if provided) filled;
                               editors complete the rest
  processing_log.json          tool version, timestamps, per-source SHA-256, transformations,
                               score-matching results
                               (KEEP PRIVATE - may contain original filenames with student names)

SCORES:
  Two ways to bring scores in, and you can use either regardless of --input type:

  1. --score-cols (xlsx --input only): additional column(s) in the SAME spreadsheet as the
     essay text, pulled in alongside it row-by-row. One column -> manifest column "score".
     Multiple columns (e.g. several raters) -> "score_1", "score_2", ... in the order given,
     matching the manifest's documented score_1/score_2/.../score_final convention (add
     score_final yourself after adjudication -- this tool doesn't compute it).

  2. --scores-file (works with either --input type): a separate CSV or XLSX with one row
     per text, a key column to match on, and one or more score columns. For directory
     --input, the key column is matched against the ORIGINAL filename (tried exact, then
     without extension, so "essay1.docx" matches either "essay1.docx" or "essay1" in the
     scores file). For xlsx --input, it's matched against --id-col's value (so give the
     same --id-col to both the essay spreadsheet and the scores file). Unmatched texts are
     left blank in the manifest and flagged in the console output and processing log for
     manual follow-up, rather than silently skipped -- not every text needs a score, but you
     should know which ones are missing.

ID REGISTRY (collision prevention across multiple submissions sharing one prefix):
  If you reuse the same --prefix across separate runs -- e.g. one institution submitting
  several exam-type subcorpora (awade-expository, awade-comparative, awade-argumentative)
  that should share one continuous ID pool ("awade-0001", "awade-0002", ... regardless of
  which submission a text belongs to) -- this tool tracks the next available number per
  prefix in a registry file (default: data/_id_registry.json) so each new run continues
  from where the last one left off instead of restarting at 0001 and colliding with an
  earlier submission's IDs. The registry is updated automatically after a successful run;
  commit it alongside the new submission. Use --start-at to override (e.g. first-ever
  seeding, or recovering from a manually-corrected collision) -- this bypasses the
  registry's stored value for this run and prints a warning.

Transformations applied (exactly these, nothing else):
  - docx: paragraph text extraction (python-docx; body paragraphs in document order)
  - decode (UTF-8 with BOM tolerance for .txt; cp932/shift-jis fallback is NOT attempted -
    undecodable files are skipped and reported so the editor resolves encoding explicitly)
  - Unicode NFC normalization
  - CRLF/CR -> LF
  - strip UTF-8 BOM; ensure single trailing newline
Word-internal characters (curly quotes, dashes, students' errors) are NOT altered.
Score values are copied verbatim as strings -- no parsing, rounding, or validation.

Requires: pip install python-docx openpyxl
"""
import argparse
import csv
import datetime
import hashlib
import json
import sys
import unicodedata
from pathlib import Path

TOOL_NAME = "intake_convert.py"
TOOL_VERSION = "1.1.0"


def normalize(text: str) -> str:
    text = text.lstrip("\ufeff")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = unicodedata.normalize("NFC", text)
    if not text.endswith("\n"):
        text += "\n"
    return text


def read_docx(path: Path) -> str:
    from docx import Document
    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs)


def read_txt(path: Path) -> str:
    return path.read_bytes().decode("utf-8")  # strict; BOM handled in normalize()


def rows_from_xlsx(path: Path, sheet: str | None, text_col: str, id_col: str | None, score_cols: list[str] | None = None):
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    if text_col not in header:
        sys.exit(f"Column '{text_col}' not found. Available: {header}")
    t_idx = header.index(text_col)
    i_idx = header.index(id_col) if id_col and id_col in header else None
    score_idxs = []
    if score_cols:
        for c in score_cols:
            if c not in header:
                sys.exit(f"--score-cols column '{c}' not found. Available: {header}")
            score_idxs.append(header.index(c))
    for n, row in enumerate(rows, start=2):
        text = row[t_idx] if t_idx < len(row) else None
        if text is None or str(text).strip() == "":
            continue
        src_id = str(row[i_idx]) if i_idx is not None and i_idx < len(row) else f"row{n}"
        scores = [str(row[i]) if i < len(row) and row[i] is not None else "" for i in score_idxs]
        yield src_id, str(text), scores


def load_external_scores(path: Path, key_col: str, value_cols: list[str]) -> dict:
    """Returns {key_value_as_string: [score1, score2, ...]}. Supports .csv and .xlsx."""
    scores = {}
    if path.suffix.lower() == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
        for col in [key_col] + value_cols:
            if col not in header:
                sys.exit(f"--scores-file column '{col}' not found. Available: {header}")
        k_idx = header.index(key_col)
        v_idxs = [header.index(c) for c in value_cols]
        for row in rows:
            key = row[k_idx]
            if key is None or str(key).strip() == "":
                continue
            scores[str(key).strip()] = [str(row[i]) if i < len(row) and row[i] is not None else "" for i in v_idxs]
    elif path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for col in [key_col] + value_cols:
                if col not in (reader.fieldnames or []):
                    sys.exit(f"--scores-file column '{col}' not found. Available: {reader.fieldnames}")
            for row in reader:
                key = row.get(key_col, "")
                if not key or not key.strip():
                    continue
                scores[key.strip()] = [row.get(c, "") for c in value_cols]
    else:
        sys.exit("--scores-file must be .csv or .xlsx")
    return scores


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def load_registry(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as ex:
        sys.exit(f"Registry {path} exists but is not valid JSON ({ex}). Fix or remove it before continuing.")


def save_registry(path: Path, registry: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(registry, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    tmp.replace(path)  # atomic on POSIX; avoids a half-written registry on interruption


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--input", required=True, help="Directory of .txt/.docx files, or an .xlsx file")
    ap.add_argument("--outdir", required=True, help="Subcorpus directory, e.g. data/kyoto-argumentative-2026")
    ap.add_argument("--prefix", required=True, help="text_id prefix, e.g. kyo26. Reuse the same prefix across multiple submissions to share one continuous ID pool (see ID REGISTRY above).")
    ap.add_argument("--text-col", help="(xlsx) header of the column containing essay text")
    ap.add_argument("--id-col", help="(xlsx) optional header of a source-identifier column (logged privately, NOT used as text_id)")
    ap.add_argument("--sheet", help="(xlsx) worksheet name; default: active sheet")
    ap.add_argument("--score-cols", help="(xlsx --input only) comma-separated column name(s) in the SAME spreadsheet holding score(s), e.g. 'rater1_score,rater2_score'. One column -> manifest 'score'; multiple -> 'score_1', 'score_2', ...")
    ap.add_argument("--scores-file", help="Path to a separate .csv or .xlsx with one row per text, for matching scores to texts that arrived as individual files (or as a spreadsheet with scores in a different file). Requires --scores-key-col and --scores-value-cols.")
    ap.add_argument("--scores-key-col", help="Column in --scores-file to match on: original filename for directory --input, or the --id-col value for xlsx --input.")
    ap.add_argument("--scores-value-cols", help="Comma-separated column name(s) in --scores-file holding the actual score(s).")
    ap.add_argument("--registry", default="data/_id_registry.json", help="Path to the ID registry JSON tracking next-available-number per prefix (default: data/_id_registry.json)")
    ap.add_argument("--start-at", type=int, help="Override the registry-derived starting number for this run (e.g. first-ever seeding, or manual collision recovery). Prints a warning; still updates the registry afterward.")
    args = ap.parse_args()

    score_cols = [c.strip() for c in args.score_cols.split(",")] if args.score_cols else []
    scores_value_cols = [c.strip() for c in args.scores_value_cols.split(",")] if args.scores_value_cols else []
    if args.scores_file and not (args.scores_key_col and scores_value_cols):
        sys.exit("--scores-file requires both --scores-key-col and --scores-value-cols")
    if score_cols and not (args.input.endswith(".xlsx")):
        sys.exit("--score-cols only applies to xlsx --input; use --scores-file for individual files")

    external_scores = {}
    unmatched_score_keys = set()
    if args.scores_file:
        external_scores = load_external_scores(Path(args.scores_file), args.scores_key_col, scores_value_cols)
        unmatched_score_keys = set(external_scores.keys())  # narrowed down as matches are found below

    total_score_cols = len(score_cols) + len(scores_value_cols)
    if total_score_cols == 0:
        score_col_names = []
    elif total_score_cols == 1:
        score_col_names = ["score"]
    else:
        score_col_names = [f"score_{i}" for i in range(1, total_score_cols + 1)]

    inp, outdir, registry_path = Path(args.input), Path(args.outdir), Path(args.registry)
    texts_dir = outdir / "texts"
    texts_dir.mkdir(parents=True, exist_ok=True)

    registry = load_registry(registry_path)
    if args.start_at is not None:
        start_n = args.start_at
        print(f"WARNING: --start-at {start_n} overrides the registry (which has "
              f"{registry.get(args.prefix, 'no entry')} for prefix '{args.prefix}'). "
              f"Make sure this doesn't collide with IDs already in use.")
    else:
        start_n = registry.get(args.prefix, 1)

    entries, skipped = [], []
    transformations = [
        "NFC normalization", "CRLF/CR -> LF", "BOM strip", "single trailing newline"
    ]

    if inp.is_dir():
        sources = sorted(p for p in inp.iterdir() if p.suffix.lower() in (".txt", ".docx") and not p.name.startswith("~$"))
        if not sources:
            sys.exit(f"No .txt or .docx files found in {inp}")
        readers = {".txt": read_txt, ".docx": read_docx}
        if any(p.suffix.lower() == ".docx" for p in sources):
            transformations.insert(0, "docx paragraph extraction (python-docx)")
        n = start_n
        for src in sources:
            tid = f"{args.prefix}-{n:04d}"
            try:
                raw = readers[src.suffix.lower()](src)
            except UnicodeDecodeError:
                skipped.append((src.name, "not UTF-8; resolve encoding explicitly and re-run"))
                continue
            except Exception as ex:  # unreadable/corrupt file - report, don't guess
                skipped.append((src.name, f"read error: {ex}"))
                continue
            out = normalize(raw)
            (texts_dir / f"{tid}.txt").write_text(out, encoding="utf-8")
            matched_scores = []
            if args.scores_file:
                key = src.name if src.name in external_scores else src.stem if src.stem in external_scores else None
                if key is not None:
                    matched_scores = external_scores[key]
                    unmatched_score_keys.discard(key)
            entries.append({
                "text_id": tid, "source_file": src.name,
                "source_sha256": sha256_file(src),
                "word_count": len(out.split()),
                "scores": matched_scores,
                "score_matched": bool(matched_scores) if args.scores_file else None,
            })
            n += 1
    elif inp.suffix.lower() == ".xlsx":
        if not args.text_col:
            sys.exit("--text-col is required for spreadsheet input")
        transformations.insert(0, "extraction from spreadsheet cells (openpyxl, values only)")
        sheet_sha = sha256_file(inp)
        n = start_n
        for src_id, raw, inline_scores in rows_from_xlsx(inp, args.sheet, args.text_col, args.id_col, score_cols):
            tid = f"{args.prefix}-{n:04d}"
            out = normalize(raw)
            (texts_dir / f"{tid}.txt").write_text(out, encoding="utf-8")
            matched_scores = list(inline_scores)
            if args.scores_file:
                if src_id in external_scores:
                    matched_scores += external_scores[src_id]
                    unmatched_score_keys.discard(src_id)
                else:
                    matched_scores += [""] * len(scores_value_cols)
            entries.append({
                "text_id": tid, "source_file": f"{inp.name}:{src_id}",
                "source_sha256": sheet_sha,
                "word_count": len(out.split()),
                "scores": matched_scores,
                "score_matched": all(matched_scores) if (score_cols or args.scores_file) else None,
            })
            n += 1
    else:
        sys.exit("--input must be a directory or an .xlsx file")

    # manifest skeleton: editors complete writer_id/genre/prompt_id/composition_date from submitter materials
    with open(outdir / "manifest_skeleton.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["text_id", "writer_id", "genre", "prompt_id", "composition_date", "word_count"] + score_col_names + ["notes"])
        for e in entries:
            scores_padded = (e["scores"] + [""] * len(score_col_names))[:len(score_col_names)]
            w.writerow([e["text_id"], "", "", "", "", e["word_count"]] + scores_padded + [""])

    log = {
        "tool": {"name": TOOL_NAME, "version": TOOL_VERSION},
        "run_at_utc": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
        "input": str(inp),
        "transformations_applied": transformations,
        "texts_converted": len(entries),
        "skipped": [{"source_file": s, "reason": r} for s, r in skipped],
        "entries": entries,
        "score_matching": {
            "scores_file_used": args.scores_file,
            "texts_missing_a_score": [e["text_id"] for e in entries if e.get("score_matched") is False],
            "scores_file_rows_never_matched": sorted(unmatched_score_keys),
        } if (args.scores_file or score_cols) else None,
        "note": "KEEP PRIVATE: source_file values may contain student-identifying original filenames.",
    }
    (outdir / "processing_log.json").write_text(json.dumps(log, indent=2), encoding="utf-8")

    if entries:
        registry[args.prefix] = start_n + len(entries)
        save_registry(registry_path, registry)

    print(f"Converted {len(entries)} text(s) -> {texts_dir}")
    if entries:
        print(f"ID range used: {entries[0]['text_id']} .. {entries[-1]['text_id']}")
        print(f"Registry updated: {registry_path} (prefix '{args.prefix}' next starts at {registry[args.prefix]:04d}) -- commit this file alongside the submission.")
    if score_col_names:
        n_missing = sum(1 for e in entries if e.get("score_matched") is False)
        print(f"Scores: {len(entries) - n_missing}/{len(entries)} text(s) matched to a score.")
        if n_missing:
            print(f"  MISSING a score: {[e['text_id'] for e in entries if e.get('score_matched') is False]}")
        if unmatched_score_keys:
            print(f"  WARNING: {len(unmatched_score_keys)} row(s) in --scores-file were never matched to any text "
                  f"(check for key typos): {sorted(unmatched_score_keys)}")
    print(f"Manifest skeleton: {outdir/'manifest_skeleton.csv'} (complete writer_id/genre/prompt_id/composition_date)")
    print(f"Processing log (PRIVATE, do not commit): {outdir/'processing_log.json'}")
    for s, r in skipped:
        print(f"SKIPPED {s}: {r}")
    return 1 if skipped else 0


if __name__ == "__main__":
    sys.exit(main())

